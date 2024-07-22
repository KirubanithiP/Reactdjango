import openpyxl
from django.http import HttpResponse
from django.core.mail import send_mail
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from backend.models import Students
from backend.serializers import StudentSerializer
from django.db.models import Q
from django.shortcuts import render
from reactdjango.settings import EMAIL_HOST_USER 
from django.conf import settings
from rest_framework.decorators import api_view
import logging
from django.core.mail import EmailMessage
import requests
from backend.models import Websites
from io import BytesIO
from django.core.files import File
from PIL import Image, ImageDraw
import qrcode
from .serializers import WebsitesSerializer
from django.shortcuts import get_object_or_404
from .models import Barcode
from django.views import View
from django.shortcuts import render, redirect
from .forms import ExcelFileForm
from django.shortcuts import render




class StudentAPI(APIView):
    def apply_filters(self, students_queryset, params):
        course = params.get('course')
        first_name = params.get('first_name')
        last_name = params.get('last_name')
        search_query = params.get('search_query')

        if course:
            students_queryset = students_queryset.filter(Course=course)
        if first_name:
            students_queryset = students_queryset.filter(Firstname=first_name)
        if last_name:
            students_queryset = students_queryset.filter(Lastname=last_name)
        if search_query:
            students_queryset = students_queryset.filter(
                Q(Firstname__icontains=search_query) |
                Q(Lastname__icontains=search_query) |
                Q(Email__icontains=search_query) |
                Q(Course__icontains=search_query)
            )
        return students_queryset

    def get(self, request, pk=None):
        try:
            if pk is not None:
                student = Students.objects.get(studentId=pk)
                serializer = StudentSerializer(student)
                edit_data = {
                    "edit": True,
                    "student": serializer.data
                }
                return Response(edit_data, status=status.HTTP_200_OK)
            else:
                students = Students.objects.all()
                filtered_students = self.apply_filters(students, request.query_params)
                serializer = StudentSerializer(filtered_students, many=True)
                return Response(serializer.data, status=status.HTTP_200_OK)
        except Students.DoesNotExist:
            return Response("Student not found", status=status.HTTP_404_NOT_FOUND)

    def post(self, request):
        try:
            student_data = request.data
            students_serializer = StudentSerializer(data=student_data)

            if not students_serializer.is_valid():
                return Response({
                    'data': students_serializer.errors,
                    'message': "Invalid student data"
                }, status=status.HTTP_400_BAD_REQUEST)

            students_serializer.save()

            return Response({"message": "Student added successfully"})
        except Exception as e:
            logging.error(f"An error occurred: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def put(self, request, pk=None):
        try:
            student = Students.objects.get(studentId=pk)
            student_data = request.data
            students_serializer = StudentSerializer(student, data=student_data)
            if students_serializer.is_valid():
                students_serializer.save()
                return Response("Updated Successfully", status=status.HTTP_200_OK)
            return Response(students_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Students.DoesNotExist:
            return Response("Student not found", status=status.HTTP_404_NOT_FOUND)

    def patch(self, request, pk=None):
        try:
            student = Students.objects.get(studentId=pk)
            serializer = StudentSerializer(student, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response("Partial Update Successful", status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Students.DoesNotExist:
            return Response("Student not found", status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk=None):
        try:
            student = Students.objects.get(studentId=pk)
            student.delete()
            return Response("Student Was Deleted Successfully", status=status.HTTP_204_NO_CONTENT)
        except Students.DoesNotExist:
            return Response("Student not found", status=status.HTTP_404_NOT_FOUND)



EMAIL_HOST_USER = 'kirubanithip2@gmail.com'



@api_view(['POST'])
def send_mail_view(request):
    try:
        data = request.data
        subject = "KIRUBANITHI"
        message = "This is a test email from your app."
        from_email = EMAIL_HOST_USER  # Use the configured sender email
        recipient_list = data.get("recipient_list", [])

        # Load the Excel data
        excel_data = fetch_data_from_backend()  # Define this function to fetch Excel data from backend
        excel_file_path = create_excel_file(excel_data)  # Define this function to create Excel file

        email = EmailMessage(subject, message, from_email, recipient_list)
        email.attach_file(excel_file_path)  # Attach the Excel file to the email

        email.send(fail_silently=False)

        logging.info("Email sent successfully")
        return Response({"message": "Email sent successfully"})
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def create_excel_file(data):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers
    headers = ["Student ID", "First Name", "Last Name", "Email", "Course"]
    sheet.append(headers)

    # Add data rows
    for row_data in data:
        row = [row_data["studentId"], row_data["Firstname"], row_data["Lastname"], row_data["Email"], row_data["Course"]]
        sheet.append(row)

    # Save the workbook to a temporary file
    temp_file_path = "temp_excel_file.xlsx"
    workbook.save(temp_file_path)

    return temp_file_path
# Fetch data from backend API endpoint
def fetch_data_from_backend():
    url = "http://127.0.0.1:8000/student/"  # Replace with your actual API endpoint
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        return data
    else:
        raise Exception("Failed to fetch data from backend")

@api_view(['GET'])
def get_excel_data(request):
    try:
        # Load the Excel file
        excel_file = openpyxl.load_workbook('students.xlsx')

        # Select the first sheet
        sheet = excel_file.active

        # Extract data from the sheet
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            student_data = {
                "studentId": row[0],
                "Firstname": row[1],
                "Lastname": row[2],
                "Email": row[3],
                "Course": row[4],
            }
            data.append(student_data)

        return Response(data, status=status.HTTP_200_OK)
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class WebsitesAPI(APIView):
    def get(self, request):
        websites = Websites.objects.all()
        serializer = WebsitesSerializer(websites, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = WebsitesSerializer(data=request.data)
        if serializer.is_valid():
            website = serializer.save()

            # Generate and save QR code for the created website
            qrcode_img = qrcode.make(website.name)
            canvas = Image.new('RGB', (290, 290), 'white')
            draw = ImageDraw.Draw(canvas)
            canvas.paste(qrcode_img)
            fname = f'qr_code-{website.name}.png'
            buffer = BytesIO()
            canvas.save(buffer, 'PNG')
            website.qr_code.save(fname, File(buffer), save=False)
            canvas.close()

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

# views.py

from .tasks import send_email_task
from datetime import datetime, timedelta

def send_email_view(request):
    # Assuming you have some logic to determine the email, subject, and message
    email = 'kirubanithip4@gmail.com'
    subject = 'Your Subject'
    message = 'Your Message'

    # Calculate the desired send time (e.g., 5 minutes from now)
    send_time = datetime.now() + timedelta(minutes=5)

    # Schedule the task to run at the desired time
    send_email_task.apply_async(args=[email, subject, message], eta=send_time)

    return render(request, 'email_sent.html')




def motech(request):
    return render(request, 'backend/home.html')






class GenerateBarcode(APIView):
    def post(self, request, format=None):
        code = request.data.get('code', '')
        if code:
            barcode, created = Barcode.objects.get_or_create(code=code)
            return Response({'barcode_id': barcode.id}, status=status.HTTP_201_CREATED)
        else:
            return Response({'error': 'Code is required.'}, status=status.HTTP_400_BAD_REQUEST)

class DisplayBarcode(APIView):
    def get(self, request, barcode_id, format=None):
        barcode = get_object_or_404(Barcode, pk=barcode_id)
        return HttpResponse(barcode.image.read(), content_type='image/png')

class BarcodeList(APIView):
    def get(self, request, format=None):
        barcodes = Barcode.objects.all()
        data = [{'id': barcode.id, 'code': barcode.code} for barcode in barcodes]
        return Response(data)


# views.py

class ExcelUploadView(View):

    def get(self, request, *args, **kwargs):
        form = ExcelFileForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = ExcelFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.save()
            return redirect('success_page', pk=excel_file.pk)
        return render(request, self.template_name, {'form': form})

class ExcelSuccessView(View):
    template_name = 'backend/success.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, {'pk': kwargs.get('pk')})
